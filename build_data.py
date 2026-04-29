#!/usr/bin/env python3
"""
Genera data.json a partir de los Excel del club.

Entradas esperadas (en SOURCE_DIR):
  - "Ordenes abonos.xlsx"  -> emisión de cada abono (una fila por asiento).
        Columnas usadas: EVENTO, FECHA, CÓDIGO DE BARRAS
  - "<Rival>.xlsx"         -> un archivo por partido con los accesos registrados.
        Columnas usadas: INGRESO, TIPO, CÓDIGO DE BARRAS

Cada archivo rival = 1 partido. La fecha del partido se toma como la fecha más
frecuente en la columna INGRESO.

Uso:
    python3 build_data.py
Salida:
    <output>/data.json
"""

import json
import os
import re
from collections import Counter
from datetime import datetime
from pathlib import Path

import openpyxl

# ---------- CONFIGURACIÓN ----------------------------------------------------
SOURCE_DIR = Path("/Users/javierolea/Desktop/Toluca")
OUTPUT     = Path("/Users/javierolea/Desktop/reporte-abonados/data.json")

ORDERS_FILE = "Ordenes abonos.xlsx"
MS_SUBDIR   = "Mercado Secundario"

# Todos los abonos son anuales esta temporada.
# Vigencia: desde la fecha de compra (no eres elegible para partidos previos a
# tu emisión) hasta el fin estimado de la Clausura 2026.
SEASON_END = "2026-05-31"

# Qué filas de accesos cuentan como "uso de abono"
ACCESO_TIPOS_VALIDOS = {"Abono", "Cortesía abono"}

# Qué filas de Ordenes son abonos reales (descarta "test", administradores, etc.)
ORDEN_EVENTO_RE = re.compile(r"DIABLO ABONO", re.IGNORECASE)
# Abonos "Digital FULL" (paquete completo): se reportan en una sección aparte.
ORDEN_FULL_RE   = re.compile(r"DIGITAL\s+FULL", re.IGNORECASE)

# Partidos con menos de N accesos de abono se consideran liguilla/no-abono y se
# excluyen (el abono regular no cubre esos partidos y todos aparecerían como
# ausentes, distorsionando el ranking). Súbelo o bájalo si hace falta.
MIN_ACCESOS_PARTIDO = 5000

# -----------------------------------------------------------------------------

def parse_date(s):
    """Los archivos vienen como 'dd/mm/yy HH:MM' o 'dd/mm/yy'. Devuelve 'YYYY-MM-DD'."""
    if s is None:
        return None
    if isinstance(s, datetime):
        return s.strftime("%Y-%m-%d")
    s = str(s).strip()
    m = re.match(r"(\d{2})/(\d{2})/(\d{2})", s)
    if not m:
        return None
    d, mo, y = m.groups()
    return f"20{y}-{mo}-{d}"


def read_sheet(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Hoja 1"] if "Hoja 1" in wb.sheetnames else wb.active
    # Fila 3 (index 2) es el encabezado real; las 2 anteriores son títulos.
    rows = ws.iter_rows(values_only=True)
    for _ in range(2):
        next(rows, None)
    header = next(rows, None)
    for r in rows:
        if r is None:
            continue
        yield dict(zip(header, r))
    wb.close()


def build_abonados():
    """Un abonado = (NÚMERO DE ORDEN, ASIENTO). El ID se compone como 'ORDEN-ASIENTO'.
    El CÓDIGO DE BARRAS de Ordenes NO sirve para ligar: el sistema reemite un
    barcode distinto por partido en los accesos. El match se hace por orden+asiento.
    Cada abonado incluye nombre, correo y teléfono del titular.
    """
    def clean(v):
        if v is None: return ""
        s = str(v).strip()
        return "" if s in ("-", "1", "0") else s

    path = SOURCE_DIR / ORDERS_FILE
    seen = {}
    total_rows = 0
    cortesias_count = 0
    for d in read_sheet(path):
        ev = d.get("EVENTO") or ""
        if not ORDEN_EVENTO_RE.search(str(ev)):
            continue
        numord  = d.get("NÚMERO DE ORDEN")
        asiento = d.get("ASIENTO")
        if not numord or not asiento:
            continue
        key = f"{numord}-{asiento}"
        fecha = parse_date(d.get("FECHA"))
        total_rows += 1
        es_full = bool(ORDEN_FULL_RE.search(str(ev)))
        # TIPO == "CORTESIA" en Ordenes marca abonos regalados.
        # No los excluimos para poder filtrarlos en el dashboard.
        es_cortesia = (str(d.get("TIPO") or "").strip().upper() == "CORTESIA")
        prev = seen.get(key)
        if prev is None or (fecha and fecha < prev["vigencia_inicio"]):
            row = {
                "id": key,
                "tipo": "anual",
                "vigencia_inicio": fecha or "2025-07-01",
                "vigencia_fin": SEASON_END,
                "nombre":   clean(d.get("*NOMBRE")),
                "correo":   clean(d.get("*CORREO ELECTRÓNICO")),
                "telefono": clean(d.get("*TELÉFONO")),
                "zona":     clean(d.get("ZONA")),
                "seccion":  clean(d.get("SECCIÓN")),
                "asiento":  str(asiento),
                "orden":    str(numord),
                "barcode":  clean(d.get("CÓDIGO DE BARRAS")),
            }
            if es_full:
                row["esFull"] = 1
            if es_cortesia:
                row["esCortesia"] = 1
                cortesias_count += 1
            seen[key] = row
    print(f"  Órdenes procesadas: {total_rows}  |  abonados únicos (orden+asiento): {len(seen)}  |  cortesías: {cortesias_count}")
    return list(seen.values())


def build_partidos_y_accesos(abonados, abonado_idx):
    """Un archivo rival por partido. Fecha = moda de INGRESO.
    Devuelve (partidos, accesos) donde accesos es {partido_id: [idx_abonado,...]}.
    Se excluyen partidos con < MIN_ACCESOS_PARTIDO (liguilla/no-abono).

    Matching por acceso:
      1. Primero intenta (COMPRA, ASIENTO) exacto.
      2. Los accesos con ASIENTO genérico (p.ej. "General") se reparten entre
         los asientos no asignados de esa misma orden.
    """
    partidos = []
    accesos  = {}
    excluidos = []

    # Asientos de cada orden (para repartir los accesos genéricos).
    orden_asientos = {}
    for i, a in enumerate(abonados):
        # a["id"] = "NUMORDEN-ASIENTO"; separa por primer guion.
        numord, _, asiento = a["id"].partition("-")
        orden_asientos.setdefault(numord, []).append((asiento, i))
    for lst in orden_asientos.values():
        lst.sort()  # orden determinístico por nombre de asiento

    # Asientos que consideramos "genéricos" (vienen sin numeración en el scanner).
    ASIENTOS_GENERICOS = {"General", "GENERAL", "general"}

    for f in sorted(SOURCE_DIR.iterdir()):
        if not f.is_file() or f.suffix.lower() != ".xlsx":
            continue
        if f.name == ORDERS_FILE:
            continue
        rival = f.stem

        fechas = Counter()
        claves_exactas  = set()        # (orden, asiento) con asiento numerado
        genericos_por_orden = Counter()  # orden -> #accesos genéricos
        for d in read_sheet(f):
            tipo = d.get("TIPO")
            ing  = parse_date(d.get("INGRESO"))
            if ing:
                fechas[ing] += 1
            if tipo not in ACCESO_TIPOS_VALIDOS:
                continue
            compra  = d.get("COMPRA")
            asiento = d.get("ASIENTO")
            if not compra:
                continue
            compra = str(compra)
            if asiento and str(asiento) not in ASIENTOS_GENERICOS:
                claves_exactas.add(f"{compra}-{asiento}")
            else:
                genericos_por_orden[compra] += 1

        if not fechas:
            print(f"  [!] {f.name}: sin fechas válidas, se omite.")
            continue

        # Total de accesos únicos (exactos + genéricos agregados)
        n_accesos = len(claves_exactas) + sum(genericos_por_orden.values())
        if n_accesos < MIN_ACCESOS_PARTIDO:
            excluidos.append((f.name, n_accesos))
            print(f"  [liguilla] {f.name:25s} accesos={n_accesos} (excluido)")
            continue

        # ----- Resolver asistencias -----
        asistieron = set()  # índices de abonados que asistieron

        # 1. Matches exactos
        match_exacto = 0
        for k in claves_exactas:
            idx = abonado_idx.get(k)
            if idx is not None:
                asistieron.add(idx)
                match_exacto += 1

        # 2. Accesos genéricos: asigna a asientos de la misma orden que aún
        #    no hayan asistido. Si hay más genéricos que asientos libres,
        #    el sobrante es "ruido" (se descarta).
        match_generico = 0
        genericos_sin_orden = 0
        for compra, n in genericos_por_orden.items():
            seats = orden_asientos.get(compra)
            if not seats:
                genericos_sin_orden += n
                continue
            libres = [idx for (_, idx) in seats if idx not in asistieron]
            asignar = min(n, len(libres))
            for idx in libres[:asignar]:
                asistieron.add(idx)
                match_generico += 1

        fecha_partido = fechas.most_common(1)[0][0]
        partido_id    = f"P{len(partidos)+1:03d}"
        partidos.append({"id": partido_id, "fecha": fecha_partido, "rival": rival})
        accesos[partido_id] = sorted(asistieron)
        print(f"  {f.name:25s} -> {fecha_partido}  accesos={n_accesos:>6}  "
              f"exacto={match_exacto}  generico={match_generico}  "
              f"huerfanos={genericos_sin_orden}")

    # Ordena partidos por fecha y reasigna IDs para que P001 = el más antiguo.
    partidos.sort(key=lambda p: p["fecha"])
    new_accesos = {}
    for i, p in enumerate(partidos, 1):
        new_id = f"P{i:03d}"
        new_accesos[new_id] = accesos[p["id"]]
        p["id"] = new_id
    return partidos, new_accesos, excluidos


def build_reventas(abonados):
    """Lee los archivos de Mercado Secundario y liga cada listing a un abonado
    por (zona, seccion, asiento). Cada archivo = un partido.
    Devuelve lista de reventas: {a: idx_abonado, rival, fecha_listado, estatus, precio}.
    """
    ms_dir = SOURCE_DIR / MS_SUBDIR
    if not ms_dir.is_dir():
        return []

    # Índice (zona, seccion, asiento) -> idx de abonado
    idx_seat = {}
    for i, a in enumerate(abonados):
        idx_seat[(a.get("zona",""), a.get("seccion",""), a.get("asiento",""))] = i

    reventas = []
    no_match = 0
    for f in sorted(ms_dir.iterdir()):
        if not f.is_file() or f.suffix.lower() != ".xlsx":
            continue
        rival = f.stem
        total_f = match_f = 0
        for d in read_sheet(f):
            estatus = d.get("ESTATUS")
            if estatus in (None, "ESTATUS"):  # skip blanks/header-dup
                continue
            total_f += 1
            k = (d.get("ZONA") or "", d.get("SECCION") or "", d.get("ASIENTO") or "")
            idx = idx_seat.get(k)
            if idx is None:
                no_match += 1
                continue
            match_f += 1
            listado = d.get("LISTADO")
            try:
                fecha = datetime.fromtimestamp(int(listado)).strftime("%Y-%m-%d") if listado else None
            except (ValueError, TypeError, OSError):
                fecha = None
            precio = d.get("PRECIO DE VENTA")
            try: precio = int(precio) if precio not in (None, "", "-") else None
            except (ValueError, TypeError): precio = None
            reventas.append({
                "a": idx,
                "rival": rival,
                "fecha": fecha,
                "estatus": "vendido" if str(estatus).upper().startswith("VEND")
                           else "expirado" if str(estatus).upper().startswith("EXP")
                           else "disponible" if str(estatus).lower().startswith("disp")
                           else str(estatus).lower(),
                "precio": precio,
            })
        print(f"  MS {rival:20s} listings={total_f:>5}  ligables={match_f}")
    print(f"  Reventas ligadas: {len(reventas)}  sin match: {no_match}")
    return reventas


def main():
    print("Leyendo abonados…")
    abonados = build_abonados()
    # Mapa barcode -> índice en el array (para compactar el JSON).
    abonado_idx = {a["id"]: i for i, a in enumerate(abonados)}

    print("Leyendo partidos y accesos…")
    partidos, accesos, excluidos = build_partidos_y_accesos(abonados, abonado_idx)

    print("Leyendo mercado secundario…")
    reventas = build_reventas(abonados)

    data = {
        "last_updated": datetime.today().strftime("%Y-%m-%d"),
        # formato compacto: accesos es {partido_id: [idx_en_abonados, ...]}
        "schema": "v2-indexed",
        "partidos": partidos,
        "abonados": abonados,
        "accesos": accesos,
        "reventas": reventas,
    }

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT, "w", encoding="utf-8") as fh:
        # sin indent para minimizar tamaño
        json.dump(data, fh, ensure_ascii=False, separators=(",", ":"))

    total_accesos = sum(len(v) for v in accesos.values())
    print(f"\nOK  -> {OUTPUT}")
    print(f"  partidos: {len(partidos)}  abonados: {len(abonados)}  accesos: {total_accesos}  reventas: {len(reventas)}")
    if excluidos:
        print("  excluidos (liguilla):", ", ".join(f"{n}({c})" for n,c in excluidos))


if __name__ == "__main__":
    main()
